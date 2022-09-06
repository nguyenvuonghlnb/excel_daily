from openpyxl import load_workbook
from datetime import datetime, date



wb = load_workbook(filename='dsc_industry_growth.xlsm')
sheet = wb.active
last_quarter = ((datetime.now().month - 1) // 3 + 1) - 1
current_year = date.today().year
print(last_quarter, current_year)
final_data_list = []

for i in range(12, sheet.max_row + 1):
    icb_level = sheet.cell(row=i, column=2).value
    if icb_level == 2:
        # rowData = []
        icb_name = sheet.cell(row=i, column=1).value
        # doanh số thuần
        net_sale = sheet.cell(row=i, column=4).value
        # tăng trưởng lãi thuần
        net_profit = sheet.cell(row=i, column=5).value
        # vốn chủ sở hữu
        equity = sheet.cell(row=i, column=6).value
        obj = {}
        obj['icbname'] = icb_name
        obj['icblevel'] = icb_level
        obj['netsale'] = net_sale
        obj['net_profit'] = net_profit
        obj['equity'] = equity
        obj['lengthreport'] = last_quarter
        obj['yearreport'] = current_year
        final_data_list.append(obj)
print(final_data_list)
