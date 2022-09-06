import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import timedelta
from openpyxl import load_workbook



list_data = []
wb = load_workbook(filename='XHTSTG.xlsm')
sheet = wb.active

data_date = sheet.cell(row=2, column=2).value
print(f'Dữ liệu ngày: {data_date} - 200')
for row_index in range(2, sheet.max_row + 1):
    data = {}
    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        if row_index > 1:
            key = sheet[column_letter + str(1)].value
            value = sheet[column_letter + str(row_index)].value
            data[key] = value
    list_data.append(data)
list_data_success = []
sub_day = 0
for x in range(0, 200 + 1):
    arr = []
    for data in list_data:
        if f'DongTien T-{x}' or f'DongTien T{x}' in data:
            obj = {}
            date = data['Date/Time'] - timedelta(days=sub_day)
            th_day = pd.Timestamp(date).dayofweek
            if th_day == 6:
                sub_day = sub_day + 2
            date = data['Date/Time'] - timedelta(days=sub_day)
            obj['asset_code'] = data['Ticker']
            obj['datetime'] = date.strftime("%Y-%m-%d %H:%M:%S")
            if x == 0:
                obj['cashflow'] = data[f'DongTien T{x}']
            else:
                obj['cashflow'] = data[f'DongTien T-{x}']
            arr.append(obj)
    sub_day = sub_day + 1
    arr = sorted(arr, key=lambda k: k['cashflow'], reverse=False)
    for n in range(0, len(arr)):
        arr[n]['top'] = n + 1
        arr[n]['ranks'] = arr[n]['top'] / len(arr) * 10
        list_data_success.append(arr[n])


