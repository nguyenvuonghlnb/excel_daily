import os
import re
import shutil
import telegram
import database
from openpyxl import load_workbook
from datetime import datetime, date


def data_xlsm_growth():
    read_not = False
    today = date.today()
    dir_path = r'\powerbi\UPLOAD_FILE'
    backup_path = r'\powerbi\Backup'
    failed_path = r'\powerbi\Failed'
    for path in os.listdir(dir_path):
        if os.path.isfile(os.path.join(dir_path, path)):
            read_not = re.search("industry_growth", path)
            if read_not:
                dir_path = f'\\powerbi\\UPLOAD_FILE\\{path}'
                break
    #####
    if read_not:
        wb = load_workbook(filename=dir_path)
        sheet = wb.active
        data_date = sheet.cell(row=9, column=2).value
        last_quarter = ((datetime.now().month - 1) // 3 + 1) - 1
        current_year = date.today().year
        # print(last_quarter, current_year)
        final_data_list = []

        for i in range(12, sheet.max_row + 1):
            icb_level = sheet.cell(row=i, column=2).value
            if icb_level == 2:
                # rowData = []
                icb_name = sheet.cell(row=i, column=1).value
                # doanh số thuần
                net_sale = sheet.cell(row=i, column=4).value
                # tăng trưởng lãi thuần
                net_profit = sheet.cell(row=i, column=5).value
                # vốn chủ sở hữu
                equity = sheet.cell(row=i, column=6).value
                obj = {}
                obj['icbname'] = icb_name
                obj['icblevel'] = icb_level
                obj['netsale'] = net_sale
                obj['net_profit'] = net_profit
                obj['equity'] = equity
                obj['lengthreport'] = last_quarter
                obj['yearreport'] = current_year
                final_data_list.append(obj)
        # print(final_data_list)
        cur = database.main.cursor()
        cur.executemany('''INSERT INTO datafeed.dsc_industry_growth_555( icbname, icblevel, netsale, net_profit, equity, 
        lengthreport, yearreport) VALUES (%s, %s, %s, %s, %s, %s, %s) ON CONFLICT ON CONSTRAINT 
        dsc_industry_growth_pk_555 DO UPDATE SET (icbname, icblevel, netsale, net_profit, equity, lengthreport, 
        yearreport) = (EXCLUDED.icbname, EXCLUDED.icblevel, EXCLUDED.netsale, EXCLUDED.net_profit, EXCLUDED.equity, 
        EXCLUDED.lengthreport, EXCLUDED.yearreport)''',
                        [[item.get("icbname"),
                          item.get("icblevel"),
                          item.get("netsale"),
                          item.get("net_profit"),
                          item.get("equity"),
                          item.get("lengthreport"),
                          item.get("yearreport")] for item in final_data_list])
        database.main.commit()
        print(f'[industry_growth] insert done data size: {len(final_data_list)}', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        print(f'[industry_growth] Data ngày: {data_date} - 200')
        telegram.send(mess=f"[industry_growth] Insert done / Data size: {len(final_data_list)}")
        shutil.move(dir_path, backup_path)
    else:
        print("[industry_growth] Không thấy file !!!")
        telegram.send(mess=f"[industry_growth] Không tìm thấy file ngày: {today}")
