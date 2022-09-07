import os
import re
import shutil
import openpyxl
import telegram
import database
from datetime import date
from openpyxl import load_workbook


def data_xlsm_market_transaction():
    read_not = False
    today = date.today()
    dir_path = '/powerbi/UPLOAD_FILE'
    backup_path = '/powerbi/Backup'
    failed_path = '/powerbi/Failed'
    dir_path_file = ''
    for path in os.listdir(dir_path):
        if os.path.isfile(os.path.join(dir_path, path)):
            read_not = re.search("transaction_day", path)
            if read_not:
                dir_path_file = f'/powerbi/UPLOAD_FILE/{path}'
                break
    ###
    if read_not:
        pxl_doc = openpyxl.load_workbook(dir_path_file)
        pxl_doc.save(f'{dir_path}/transaction_day_{today}.xlsx')
        os.remove(dir_path_file)
        final_data_list = []
        wb = load_workbook(f'{dir_path}/transaction_day_{today}.xlsx')
        sheet = wb.active
        date_str = sheet.cell(row=10, column=2).value.split('/')[1]
        if date_str == today.strftime("%Y-%m-%d"):
            for i in range(16, sheet.max_row + 1):
                data_collection_date = sheet.cell(row=10, column=2).value
                date_date = sheet.cell(row=i, column=1).value
                if date_date is None:
                    break
                final_data_list.append({
                    "date": date_date,
                    "volume_buy": sheet.cell(row=i, column=2).value,
                    "value_buy": sheet.cell(row=i, column=3).value,
                    "volume_sell": sheet.cell(row=i, column=4).value,
                    "value_sell": sheet.cell(row=i, column=5).value,
                    "volume_net": sheet.cell(row=i, column=6).value,
                    "value_net": sheet.cell(row=i, column=7).value,
                    "data_collection_date": data_collection_date,
                    "category": "individual_domestic"
                })
                final_data_list.append({
                    "date": date_date,
                    "volume_buy": sheet.cell(row=i, column=8).value,
                    "value_buy": sheet.cell(row=i, column=9).value,
                    "volume_sell": sheet.cell(row=i, column=10).value,
                    "value_sell": sheet.cell(row=i, column=11).value,
                    "volume_net": sheet.cell(row=i, column=12).value,
                    "value_net": sheet.cell(row=i, column=13).value,
                    "data_collection_date": data_collection_date,
                    "category": "individuals_foreign"
                })
                final_data_list.append({
                    "date": date_date,
                    "volume_buy": sheet.cell(row=i, column=14).value,
                    "value_buy": sheet.cell(row=i, column=15).value,
                    "volume_sell": sheet.cell(row=i, column=16).value,
                    "value_sell": sheet.cell(row=i, column=17).value,
                    "volume_net": sheet.cell(row=i, column=18).value,
                    "value_net": sheet.cell(row=i, column=19).value,
                    "data_collection_date": data_collection_date,
                    "category": "organization_domestic"
                })
                final_data_list.append({
                    "date": date_date,
                    "volume_buy": sheet.cell(row=i, column=20).value,
                    "value_buy": sheet.cell(row=i, column=21).value,
                    "volume_sell": sheet.cell(row=i, column=22).value,
                    "value_sell": sheet.cell(row=i, column=23).value,
                    "volume_net": sheet.cell(row=i, column=24).value,
                    "value_net": sheet.cell(row=i, column=25).value,
                    "data_collection_date": data_collection_date,
                    "category": "organization_foreign"
                })

            cur = database.main.cursor()
            cur.executemany('''INSERT INTO datafeed.dsc_market_transaction_net (category, date, volume_buy, value_buy, 
            volume_sell, value_sell, volume_net, value_net, data_collection_date) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 
            %s) ON CONFLICT ON CONSTRAINT dsc_mst_asset_transaction_statistics_general_pkey DO UPDATE SET (category, 
            date, volume_buy, value_buy, volume_sell, value_sell, volume_net, value_net, data_collection_date) = (
            EXCLUDED.category, EXCLUDED.date, EXCLUDED.volume_buy, EXCLUDED.value_buy, EXCLUDED.volume_sell, 
            EXCLUDED.value_sell, EXCLUDED.volume_net, EXCLUDED.value_net, EXCLUDED.data_collection_date)''',
                            [[
                                item.get("category"),
                                item.get("date"),
                                item.get("volume_buy"),
                                item.get("value_buy"),
                                item.get("volume_sell"),
                                item.get("value_sell"),
                                item.get("volume_net"),
                                item.get("value_net"),
                                item.get("data_collection_date")
                            ] for item in final_data_list])
            print(f'[transaction_day/daily] Insert done / Data size: {len(final_data_list)} - Day: {date_str}')
            telegram.send(mess=f"[transaction_day/daily] Insert done / Data size: {len(final_data_list)} - Day: {date_str}")
            database.main.commit()
            shutil.move(f'{dir_path}/transaction_day_{today}.xlsx', backup_path)
        else:
            shutil.move(f'{dir_path}/transaction_day_{today}.xlsx', failed_path)
            print(f'[transaction_day/daily] Error: Not date data: {today}. Data date: {date_str}')
            telegram.send(mess=f"[transaction_day/daily] Error: Wrong data. Data date: {date_str}")
        read_not = False
    else:
        print(f"[transaction_day/daily] Caution: File not found / Date: {today}")
        telegram.send(mess=f"[transaction_day/daily] Caution: File not found / Date: {today}")


