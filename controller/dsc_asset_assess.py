import os
import re
import shutil
import telegram
import database
from openpyxl import load_workbook
from datetime import date


def data_xlsx_backtest():
    read_not = False
    today = date.today()
    dir_path = r'\powerbi\UPLOAD_FILE'
    backup_path = r'\powerbi\Backup'
    failed_path = r'\powerbi\Failed'
    for path in os.listdir(dir_path):
        if os.path.isfile(os.path.join(dir_path, path)):
            read_not = re.search("backtest", path)
            if read_not:
                dir_path = f'\\powerbi\\UPLOAD_FILE\\{path}'
                break
    #####
    if read_not:
        final_data_list = []
        wb = load_workbook(filename=dir_path)
        sheet = wb.active
        time = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
        growth = [sheet.cell(row=i, column=2).value for i in range(2, sheet.max_row + 1)]
        momentum = [sheet.cell(row=i, column=3).value for i in range(2, sheet.max_row + 1)]
        profit_max = [sheet.cell(row=i, column=4).value for i in range(2, sheet.max_row + 1)]
        profit_min = [sheet.cell(row=i, column=5).value for i in range(2, sheet.max_row + 1)]
        profit_average = [sheet.cell(row=i, column=6).value for i in range(2, sheet.max_row + 1)]

        for i in range(0, len(time)):
            assess = {}
            assess['time'] = time[i]
            assess['growth'] = growth[i]
            assess['momentum'] = momentum[i]
            assess['profit_max'] = profit_max[i]
            assess['profit_min'] = profit_min[i]
            assess['profit_average'] = profit_average[i]
            final_data_list.append(assess)
        cur = database.main.cursor()
        cur.executemany('''INSERT INTO datafeed.dsc_asset_assess_555(time, growth, momentum, profit_max, profit_min,
        profit_average) VALUES (%s, %s, %s, %s, %s, %s) ON CONFLICT ON CONSTRAINT dsc_mst_asset_assess_pkey_555 
        DO UPDATE SET (time, growth, momentum, profit_max, profit_min,profit_average) = 
        (EXCLUDED.time, EXCLUDED.growth, EXCLUDED.momentum, EXCLUDED.profit_max, EXCLUDED.profit_min, EXCLUDED.profit_average)''',
                        [[item.get("time"),
                          item.get("growth"),
                          item.get("momentum"),
                          item.get("profit_max"),
                          item.get("profit_min"),
                          item.get("profit_average")] for item in final_data_list])
        database.main.commit()
        print(f'[backtest] insert done {len(final_data_list)}')
        telegram.send(mess=f"[backtest] Insert done / Data size: {len(final_data_list)}")
        shutil.move(dir_path, backup_path)
    else:
        telegram.send(mess=f"[backtest] Không tìm thấy file ngày: {today}")
        print("[backtest] Không thấy file !!!")