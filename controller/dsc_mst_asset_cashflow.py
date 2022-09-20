import re
import os
import shutil
import telegram
import database
import pandas as pd
from datetime import datetime, timedelta, date
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


def data_xlsm():
    read_not = False
    today = date.today()
    dir_path = '/powerbi/UPLOAD_FILE'
    backup_path = '/powerbi/Backup'
    failed_path = '/powerbi/Failed'
    for path in os.listdir(dir_path):
        if os.path.isfile(os.path.join(dir_path, path)):
            read_not = re.search("world_assets", path)
            if read_not:
                dir_path = f'/powerbi/UPLOAD_FILE/{path}'
                break
    ###
    if read_not:
        list_data = []
        wb = load_workbook(filename=dir_path)
        sheet = wb.active
        data_date = sheet.cell(row=2, column=2).value
        date_format = data_date.strftime("%m/%d/%Y")
        if date_format == today.strftime("%m/%d/%Y"):
            for row_index in range(2, sheet.max_row + 1):
                data = {}
                for column_index in range(1, sheet.max_column + 1):
                    column_letter = get_column_letter(column_index)
                    if row_index > 1:
                        key = sheet[column_letter + str(1)].value
                        value = sheet[column_letter + str(row_index)].value
                        data[key] = value
                list_data.append(data)
            list_data_success = []
            sub_day = 0
            qk2 = datetime(2022, 9, 2, 00, 00, 00)
            td = datetime(2022, 1, 3, 00, 00, 00)
            ta5 = datetime(2022, 2, 4, 00, 00, 00)
            gthv = datetime(2022, 4, 11, 00, 00, 00)
            qtld2 = datetime(2022, 5, 3, 00, 00, 00)
            for x in range(0, 200 + 1):
                arr = []
                for data in list_data:
                    if f'DongTien T-{x}' or f'DongTien T{x}' in data:
                        obj = {}
                        date_time = data['Date/Time'] - timedelta(days=sub_day)
                        th_day = pd.Timestamp(date_time).dayofweek
                        if th_day == 6: sub_day = sub_day + 2
                        date_time = data['Date/Time'] - timedelta(days=sub_day)
                        if date_time == qk2: sub_day = sub_day + 2
                        if date_time == td: sub_day = sub_day + 3
                        if date_time == ta5: sub_day = sub_day + 7
                        if date_time == gthv: sub_day = sub_day + 3
                        if date_time == qtld2: sub_day = sub_day + 4

                        date_time = data['Date/Time'] - timedelta(days=sub_day)
                        obj['asset_code'] = data['Ticker']
                        obj['datetime'] = date_time.strftime("%Y-%m-%d %H:%M:%S")
                        if x == 0:
                            obj['cashflow'] = data[f'DongTien T{x}']
                        else:
                            obj['cashflow'] = data[f'DongTien T-{x}']
                        arr.append(obj)
                sub_day = sub_day + 1
                arr = sorted(arr, key=lambda k: k['cashflow'], reverse=False)
                for n in range(0, len(arr)):
                    arr[n]['top'] = n + 1
                    arr[n]['ranks'] = arr[n]['top'] / len(arr) * 10
                    list_data_success.append(arr[n])

            # for get in list_data_success:
            cur = database.main.cursor()
            cur.executemany('''INSERT INTO datafeed.dsc_mst_asset_cashflow (asset_code, datetime, cashflow, 
            ranks) VALUES (%s, %s, %s, %s)ON CONFLICT ON CONSTRAINT dsc_mst_asset_cashflow_pkey  DO UPDATE SET (
            asset_code, datetime, cashflow, ranks) = (EXCLUDED.asset_code, EXCLUDED.datetime, EXCLUDED.cashflow, 
            EXCLUDED.ranks)''',
                            [[item.get("asset_code"),
                              item.get("datetime"),
                              item.get("cashflow"),
                              item.get("ranks")] for item in list_data_success])
            database.main.commit()
            print(f'[world_assets/daily] Insert done data size: {len(list_data_success)} - Day: {data_date}', )
            telegram.send(mess=f"[world_assets/daily] Insert done / Data size: {len(list_data_success)} - Day: {data_date} ")
            shutil.move(dir_path, backup_path)
        else:
            shutil.move(dir_path, failed_path)
            print(f'[world_assets/daily] Error: Not date data: {today}. Data date: {data_date}')
            telegram.send(mess=f"[world_assets/daily] Error: Wrong data. Data date: {data_date}")
        read_not = False
    else:
        print(f"[world_assets/daily] Caution: File not found / Date: {today}")
        telegram.send(mess=f"[world_assets/daily] Caution: File not found / Date: {today}")